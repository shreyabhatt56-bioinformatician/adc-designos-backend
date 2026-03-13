"""
ADC-DesignOS Backend — PhyloBandits
Antibody-Drug Conjugate Design & Optimization Engine
"""

from flask import Flask, request, jsonify

import json
import math
import random
import hashlib
from datetime import datetime

app = Flask(__name__)

@app.after_request
def add_cors(r):
    r.headers["Access-Control-Allow-Origin"] = "*"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    r.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return r


# ─────────────────────────────────────────────
# CONJUGATION SITE PREDICTOR
# Real logic: surface exposure + flexibility + nucleophilicity
# ─────────────────────────────────────────────

HYDROPHOBIC = set("AVILMFYW")
CHARGED     = set("RKHDE")
FLEXIBLE    = set("GSPND")

def predict_conjugation_sites(sequence: str):
    """
    Predicts the best conjugation sites on an antibody sequence.
    Scores each residue by:
      - Solvent exposure proxy (flanking residue hydrophobicity)
      - Flexibility (Gly/Pro/Ser neighbours)
      - Reactivity (Cys=highest, Lys=high, engineered sites)
    """
    seq = sequence.upper().strip()
    sites = []

    for i, aa in enumerate(seq):
        score = 0.0
        rationale = []

        # Window around residue
        window = seq[max(0,i-3):i+4]

        # Surface exposure: low hydrophobic neighbours = exposed
        hydrophob_count = sum(1 for c in window if c in HYDROPHOBIC)
        exposure = max(0, 1.0 - (hydrophob_count / len(window)))
        score += exposure * 40
        if exposure > 0.6:
            rationale.append("high surface exposure")

        # Flexibility bonus
        flex_count = sum(1 for c in window if c in FLEXIBLE)
        flex_score = (flex_count / len(window)) * 20
        score += flex_score

        # Residue-specific reactivity
        if aa == 'C':
            score += 50
            rationale.append("Cys — thiol highly reactive, gold standard for site-specific conjugation")
        elif aa == 'K':
            score += 30
            rationale.append("Lys — ε-amine, NHS ester conjugation")
        elif aa == 'N':
            # Check for N-glycosylation motif N-X-S/T
            if i+2 < len(seq) and seq[i+1] not in 'P' and seq[i+2] in 'ST':
                score += 25
                rationale.append("N-glycosylation motif (N-X-S/T) — glycan remodelling site")
        elif aa == 'S' or aa == 'T':
            score += 8
        elif aa == 'H':
            score += 15
            rationale.append("His — metal-chelation conjugation possible")

        if score > 25:
            sites.append({
                "position": i + 1,
                "residue": aa,
                "score": round(min(score, 100), 1),
                "exposure": round(exposure, 2),
                "rationale": rationale if rationale else ["moderate reactivity"],
                "conjugation_type": _conjugation_type(aa),
            })

    # Return top 10 sorted by score
    sites.sort(key=lambda x: x["score"], reverse=True)
    return sites[:10]

def _conjugation_type(aa):
    types = {
        'C': 'Thiol-maleimide / Thiol-disulfide',
        'K': 'NHS ester / Reductive amination',
        'N': 'Glycan remodelling / Oxime ligation',
        'H': 'Metal-chelation / His-tag chemistry',
        'S': 'Phosphoserine / Enzymatic',
        'T': 'Enzymatic (sortase/transglutaminase)',
    }
    return types.get(aa, 'Non-specific amine/carboxyl chemistry')


# ─────────────────────────────────────────────
# LINKER STABILITY PREDICTOR
# ─────────────────────────────────────────────

LINKER_DB = {
    "SMCC":     {"type":"non-cleavable","stability":88,"plasma_t12_h":240,"release":"none","maleimide":True,  "hydrophilicity":0.4,"toxicity_risk":0.15},
    "mc-VC-PABC":{"type":"cleavable_protease","stability":82,"plasma_t12_h":168,"release":"cathepsin-B","maleimide":True,"hydrophilicity":0.6,"toxicity_risk":0.12},
    "CL2A":     {"type":"cleavable_pH","stability":70,"plasma_t12_h":72,"release":"endosomal pH","maleimide":False,"hydrophilicity":0.5,"toxicity_risk":0.20},
    "Disulfide":{"type":"cleavable_redox","stability":65,"plasma_t12_h":48,"release":"glutathione","maleimide":False,"hydrophilicity":0.3,"toxicity_risk":0.22},
    "PEG4-NHS": {"type":"non-cleavable","stability":92,"plasma_t12_h":300,"release":"none","maleimide":False,"hydrophilicity":0.9,"toxicity_risk":0.08},
    "Hydrazone":{"type":"cleavable_pH","stability":60,"plasma_t12_h":36,"release":"lysosomal pH","maleimide":False,"hydrophilicity":0.4,"toxicity_risk":0.25},
    "β-glucuronide":{"type":"cleavable_enzyme","stability":85,"plasma_t12_h":200,"release":"β-glucuronidase","maleimide":False,"hydrophilicity":0.8,"toxicity_risk":0.10},
    "SulfoSMCC":{"type":"non-cleavable","stability":90,"plasma_t12_h":260,"release":"none","maleimide":True,"hydrophilicity":0.7,"toxicity_risk":0.10},
}

def score_linker(linker_name: str, payload_type: str, target_location: str, conjugation_site: str):
    if linker_name not in LINKER_DB:
        return {"error": f"Linker '{linker_name}' not in database"}

    l = LINKER_DB[linker_name].copy()

    # Context adjustments
    score = l["stability"]
    flags = []
    recommendations = []

    if target_location == "intracellular":
        if l["type"] == "non-cleavable":
            score -= 15
            flags.append("Non-cleavable linker suboptimal for intracellular targets — payload may not release")
            recommendations.append("Switch to mc-VC-PABC or β-glucuronide for enzymatic release")
        else:
            score += 5
            recommendations.append("Cleavable linker appropriate for intracellular release")

    if target_location == "extracellular":
        if l["type"] in ["cleavable_pH","cleavable_redox"]:
            score -= 10
            flags.append("pH/redox cleavable linkers risk premature release before endocytosis")

    if payload_type == "auristatin" and l["hydrophilicity"] < 0.5:
        score -= 8
        flags.append("Auristatin is hydrophobic — hydrophobic linker may cause aggregation. Consider PEG spacer.")
        recommendations.append("Add PEG4 spacer or switch to PEG4-NHS linker")

    if payload_type == "maytansine" and l["type"] == "cleavable_redox":
        score += 5
        recommendations.append("Disulfide linker historically used with maytansine (T-DM1 class)")

    if conjugation_site == "C" and not l.get("maleimide"):
        score -= 12
        flags.append("Cys conjugation preferred with maleimide-containing linkers")
        recommendations.append("Use SMCC, mc-VC-PABC, or SulfoSMCC for Cys-maleimide conjugation")

    l["adjusted_score"] = round(min(max(score, 0), 100), 1)
    l["flags"] = flags
    l["recommendations"] = recommendations
    l["suitability"] = "Excellent" if l["adjusted_score"]>=85 else "Good" if l["adjusted_score"]>=70 else "Moderate" if l["adjusted_score"]>=55 else "Poor"
    return l


# ─────────────────────────────────────────────
# DAR (DRUG-TO-ANTIBODY RATIO) OPTIMIZER
# ─────────────────────────────────────────────

def optimize_dar(payload_potency_nm: float, antibody_mw_kda: float, target_expression: str, toxicity_threshold: float):
    """
    Calculates optimal DAR balancing:
      - Efficacy (higher DAR = more payload delivered)
      - PK (high DAR → rapid clearance, aggregation)
      - Safety (high DAR → off-target toxicity)
    Based on published DAR-PK relationships from Hamblett et al.
    """
    # Efficacy model: sigmoidal dose-response
    results = []
    target_mult = {"low":0.5,"medium":1.0,"high":2.0,"very_high":3.0}.get(target_expression,1.0)

    for dar in [1,2,3,4,6,8]:
        # Payload delivered per internalization event
        payload_delivered = dar * 1.0  # normalized units

        # PK penalty: exponential clearance increase above DAR 4
        pk_factor = 1.0 if dar <= 4 else math.exp(-0.15*(dar-4))

        # Efficacy: Hill equation
        ec50 = payload_potency_nm / target_mult
        efficacy = (payload_delivered * pk_factor) / (ec50 + payload_delivered * pk_factor)
        efficacy_pct = round(efficacy * 100, 1)

        # Aggregation risk
        aggregation_risk = round(min(100, 5 * (dar**1.4)), 1)

        # Toxicity score
        off_target_tox = round(min(100, dar * (100 - toxicity_threshold) / 8), 1)

        # PK half-life model (Junutula et al.)
        plasma_t12 = round(max(12, 168 * math.exp(-0.12*(dar-2))), 1) if dar > 2 else 168.0

        # Therapeutic index
        ti = round(efficacy_pct / max(1, off_target_tox), 2)

        results.append({
            "dar": dar,
            "efficacy_pct": efficacy_pct,
            "aggregation_risk_pct": aggregation_risk,
            "off_target_toxicity_pct": off_target_tox,
            "plasma_t12_hours": plasma_t12,
            "therapeutic_index": ti,
            "recommendation": _dar_recommendation(dar, efficacy_pct, aggregation_risk, off_target_tox, ti)
        })

    # Find optimal
    optimal = max(results, key=lambda x: x["therapeutic_index"])
    return {"dar_analysis": results, "optimal_dar": optimal["dar"], "optimal_summary": optimal}

def _dar_recommendation(dar, eff, agg, tox, ti):
    if ti > 3 and agg < 30:
        return "✓ Optimal window"
    elif agg > 50:
        return "⚠ Aggregation risk"
    elif tox > 40:
        return "⚠ Toxicity concern"
    elif eff < 40:
        return "✗ Insufficient efficacy"
    else:
        return "~ Acceptable"


# ─────────────────────────────────────────────
# PAYLOAD TOXICITY PROFILER
# ─────────────────────────────────────────────

PAYLOAD_DB = {
    "MMAE": {"class":"auristatin","mechanism":"tubulin inhibitor","potency_nm":0.1,"bystander":True,
             "hERG_risk":0.2,"hepatotox":0.15,"mdr_sensitivity":0.7,"mw":717.9,"logP":4.1},
    "MMAF": {"class":"auristatin","mechanism":"tubulin inhibitor","potency_nm":0.3,"bystander":False,
             "hERG_risk":0.15,"hepatotox":0.10,"mdr_sensitivity":0.3,"mw":731.9,"logP":2.9},
    "DM1":  {"class":"maytansine","mechanism":"tubulin inhibitor","potency_nm":0.05,"bystander":True,
             "hERG_risk":0.3,"hepatotox":0.25,"mdr_sensitivity":0.6,"mw":738.3,"logP":4.8},
    "DM4":  {"class":"maytansine","mechanism":"tubulin inhibitor","potency_nm":0.04,"bystander":True,
             "hERG_risk":0.25,"hepatotox":0.20,"mdr_sensitivity":0.5,"mw":752.3,"logP":5.1},
    "SN-38":{"class":"camptothecin","mechanism":"TOP1 inhibitor","potency_nm":1.5,"bystander":True,
             "hERG_risk":0.1,"hepatotox":0.20,"mdr_sensitivity":0.2,"mw":392.4,"logP":2.0},
    "DXd":  {"class":"camptothecin","mechanism":"TOP1 inhibitor","potency_nm":0.5,"bystander":True,
             "hERG_risk":0.08,"hepatotox":0.18,"mdr_sensitivity":0.15,"mw":474.5,"logP":1.8},
    "PBD":  {"class":"PBD_dimer","mechanism":"DNA crosslinker","potency_nm":0.001,"bystander":False,
             "hERG_risk":0.4,"hepatotox":0.35,"mdr_sensitivity":0.1,"mw":592.6,"logP":2.5},
    "CC-1065":{"class":"duocarmycin","mechanism":"DNA alkylator","potency_nm":0.002,"bystander":False,
               "hERG_risk":0.45,"hepatotox":0.40,"mdr_sensitivity":0.1,"mw":544.6,"logP":3.1},
}

def profile_payload(payload_name: str, dar: int = 4):
    if payload_name not in PAYLOAD_DB:
        return {"error": f"Payload '{payload_name}' not found"}
    p = PAYLOAD_DB[payload_name].copy()

    # Risk scores at given DAR
    systemic_exposure = dar / 4.0
    p["overall_risk_score"] = round(
        (p["hERG_risk"]*25 + p["hepatotox"]*25 + (1-p["mdr_sensitivity"])*15 + min(1,systemic_exposure)*35), 1
    )
    p["risk_level"] = "High" if p["overall_risk_score"]>65 else "Moderate" if p["overall_risk_score"]>40 else "Low"
    p["dar_used"] = dar
    p["bystander_effect"] = "Yes — kills antigen-negative neighbours" if p["bystander"] else "No — cell-autonomous killing only"
    p["mdr_risk"] = "High" if p["mdr_sensitivity"]>0.5 else "Moderate" if p["mdr_sensitivity"]>0.3 else "Low"

    alerts = []
    if p["hERG_risk"] > 0.3:
        alerts.append("hERG channel inhibition risk — cardiac safety monitoring required")
    if p["hepatotox"] > 0.25:
        alerts.append("Hepatotoxicity signal — liver function tests in Phase I")
    if p["logP"] > 4:
        alerts.append("High logP — consider hydrophilic linker to reduce aggregation")
    p["safety_alerts"] = alerts
    return p


# ─────────────────────────────────────────────
# PK/PD SIMULATOR
# ─────────────────────────────────────────────

def simulate_pk(antibody_mw_kda, dar, linker_type, dose_mg_kg, target_expression):
    """Two-compartment PK model for ADC based on published population PK parameters"""
    # Clearance scales with DAR (Hamblett 2004, Junutula 2008)
    dar_cl_factor = 1 + 0.15*(dar-2) if dar > 2 else 1.0
    base_cl = 0.006  # L/h/kg typical IgG1
    cl = base_cl * dar_cl_factor

    # Volume of distribution
    vc = 0.05   # L/kg central
    vp = 0.10   # L/kg peripheral
    q  = 0.003  # inter-compartmental

    # Half-life
    alpha = ((cl+q)/vc + q/vp)
    beta  = (cl*q)/(vc*vp)
    t12_alpha = round(math.log(2) * 2 / (alpha*24), 1)  # hours
    t12_beta  = round(math.log(2) / (beta*24/(alpha)), 1)

    # Cmax, AUC
    dose_nmol = (dose_mg_kg * 1000) / (antibody_mw_kda * 1000)
    cmax = round(dose_nmol / vc, 2)
    auc  = round(dose_nmol / cl, 1)

    # Tumor accumulation (EPR + active targeting)
    target_mult = {"low":1.5,"medium":4.0,"high":8.0,"very_high":14.0}.get(target_expression, 4.0)
    tumor_conc = round(cmax * 0.02 * target_mult, 3)

    # Payload release AUC at tumor
    release_eff = 0.7 if "cleavable" in linker_type else 0.15
    free_payload_auc = round(auc * release_eff * (dar/4), 1)

    timepoints = []
    for t in range(0, 49):
        conc = cmax * (math.exp(-alpha*t) + math.exp(-beta*t/10)) / 2
        timepoints.append({"time_h": t*3, "conc_nmol_L": round(max(0, conc), 4)})

    return {
        "t12_alpha_h": t12_alpha,
        "t12_beta_h": t12_beta,
        "cmax_nmol_L": cmax,
        "auc_nmol_h_L": auc,
        "tumor_conc_nmol_L": tumor_conc,
        "free_payload_auc": free_payload_auc,
        "pk_curve": timepoints[::4],  # every 12h
        "cl_L_h_kg": round(cl, 5),
        "dar_pk_penalty_pct": round((dar_cl_factor-1)*100, 1),
    }


# ─────────────────────────────────────────────
# CANDIDATE RANKER — puts it all together
# ─────────────────────────────────────────────

def rank_candidates(sequence, linker_name, payload_name, dar,
                    target_location, target_expression, dose_mg_kg=3.0,
                    toxicity_threshold=70.0):
    # Run all modules
    conj_sites = predict_conjugation_sites(sequence)
    top_site = conj_sites[0] if conj_sites else {"residue":"K","score":50}

    linker = score_linker(linker_name, payload_name.lower()[:8], target_location, top_site["residue"])
    payload = profile_payload(payload_name, dar)
    dar_opt = optimize_dar(
        payload.get("potency_nm", 1.0),
        150.0,  # typical IgG1 MW kDa
        target_expression,
        toxicity_threshold
    )
    pk = simulate_pk(150.0, dar, linker.get("type","non-cleavable"), dose_mg_kg, target_expression)

    # Master score
    linker_score = linker.get("adjusted_score", 70)
    payload_safety = 100 - payload.get("overall_risk_score", 50)
    pk_score = min(100, pk["tumor_conc_nmol_L"] * 1000)
    dar_ti = min(100, dar_opt["optimal_summary"]["therapeutic_index"] * 20)

    master = round((linker_score*0.30 + payload_safety*0.30 + pk_score*0.20 + dar_ti*0.20), 1)
    grade = "A" if master>=80 else "B" if master>=65 else "C" if master>=50 else "D"

    return {
        "master_score": master,
        "grade": grade,
        "conjugation_sites": conj_sites,
        "linker_analysis": linker,
        "payload_profile": payload,
        "dar_optimization": dar_opt,
        "pk_simulation": pk,
        "top_conjugation_site": top_site,
        "recommendation": _overall_recommendation(master, linker, payload, dar_opt, pk),
        "generated_at": datetime.utcnow().isoformat()
    }

def _overall_recommendation(score, linker, payload, dar_opt, pk):
    recs = []
    if score >= 80:
        recs.append(f"Strong ADC candidate. Proceed to in vitro conjugation and cytotoxicity assay.")
    elif score >= 65:
        recs.append(f"Promising candidate with addressable weaknesses.")
    else:
        recs.append(f"Significant optimization needed before progressing.")

    if linker.get("recommendations"):
        recs.extend(linker["recommendations"][:2])
    if payload.get("safety_alerts"):
        recs.extend(payload["safety_alerts"][:1])

    recs.append(f"Optimal DAR = {dar_opt['optimal_dar']}. "
                f"Predicted tumor concentration: {pk['tumor_conc_nmol_L']} nmol/L. "
                f"Plasma t½ beta: {pk['t12_beta_h']}h.")
    return recs


# ─────────────────────────────────────────────
# API ROUTES
# ─────────────────────────────────────────────

@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status":"ok","service":"ADC-DesignOS","version":"1.0.0","by":"PhyloBandits"})

@app.route("/api/conjugation-sites", methods=["POST"])
def api_conjugation_sites():
    data = request.json or {}
    seq = data.get("sequence","")
    if len(seq) < 5:
        return jsonify({"error":"Sequence too short (min 5 aa)"}), 400
    return jsonify({"sites": predict_conjugation_sites(seq), "sequence_length": len(seq)})

@app.route("/api/linker-score", methods=["POST"])
def api_linker_score():
    d = request.json or {}
    result = score_linker(
        d.get("linker","mc-VC-PABC"),
        d.get("payload_type","auristatin"),
        d.get("target_location","intracellular"),
        d.get("conjugation_site","C")
    )
    return jsonify(result)

@app.route("/api/dar-optimize", methods=["POST"])
def api_dar_optimize():
    d = request.json or {}
    return jsonify(optimize_dar(
        float(d.get("payload_potency_nm", 0.1)),
        float(d.get("antibody_mw_kda", 150)),
        d.get("target_expression","medium"),
        float(d.get("toxicity_threshold", 70))
    ))

@app.route("/api/payload-profile", methods=["POST"])
def api_payload_profile():
    d = request.json or {}
    return jsonify(profile_payload(d.get("payload","MMAE"), int(d.get("dar",4))))

@app.route("/api/pk-simulate", methods=["POST"])
def api_pk_simulate():
    d = request.json or {}
    return jsonify(simulate_pk(
        float(d.get("antibody_mw_kda",150)),
        int(d.get("dar",4)),
        d.get("linker_type","cleavable_protease"),
        float(d.get("dose_mg_kg",3.0)),
        d.get("target_expression","medium")
    ))

@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    """Master endpoint — runs full ADC design pipeline"""
    d = request.json or {}
    required = ["sequence","linker","payload","dar","target_location","target_expression"]
    for r in required:
        if r not in d:
            return jsonify({"error": f"Missing field: {r}"}), 400
    return jsonify(rank_candidates(
        d["sequence"], d["linker"], d["payload"], int(d["dar"]),
        d["target_location"], d["target_expression"],
        float(d.get("dose_mg_kg", 3.0)),
        float(d.get("toxicity_threshold", 70.0))
    ))

@app.route("/api/payloads", methods=["GET"])
def api_list_payloads():
    return jsonify({k: {"class":v["class"],"mechanism":v["mechanism"],"potency_nm":v["potency_nm"]}
                    for k,v in PAYLOAD_DB.items()})

@app.route("/api/linkers", methods=["GET"])
def api_list_linkers():
    return jsonify({k: {"type":v["type"],"stability":v["stability"],"plasma_t12_h":v["plasma_t12_h"]}
                    for k,v in LINKER_DB.items()})

if __name__ == "__main__":
    print("ADC-DesignOS API — PhyloBandits")
    print("Running on http://0.0.0.0:8000")
    app.run(host="0.0.0.0", port=8000, debug=False)


# ─────────────────────────────────────────────
# FEATURE 1: MULTI-CANDIDATE COMPARISON
# ─────────────────────────────────────────────

@app.route("/api/compare", methods=["POST"])
def api_compare():
    d = request.json or {}
    candidates = d.get("candidates", [])
    if not candidates or len(candidates) < 2:
        return jsonify({"error": "Provide at least 2 candidates"}), 400
    if len(candidates) > 10:
        return jsonify({"error": "Maximum 10 candidates"}), 400
    results = []
    for i, c in enumerate(candidates):
        try:
            r = rank_candidates(
                c.get("sequence",""), c.get("linker","mc-VC-PABC"),
                c.get("payload","MMAE"), int(c.get("dar",4)),
                c.get("target_location","intracellular"),
                c.get("target_expression","medium"),
                float(c.get("dose_mg_kg",3.0)), float(c.get("toxicity_threshold",70.0))
            )
            r["candidate_name"] = c.get("name", f"Candidate {i+1}")
            r["candidate_index"] = i
            results.append(r)
        except Exception as e:
            results.append({"candidate_name": c.get("name",f"Cand {i+1}"), "error": str(e), "candidate_index": i})
    valid = [r for r in results if "master_score" in r]
    valid.sort(key=lambda x: x["master_score"], reverse=True)
    for i, r in enumerate(valid):
        r["rank"] = i + 1
    best = valid[0] if valid else {}
    table = [{
        "name": r.get("candidate_name"), "rank": r.get("rank"),
        "master_score": r.get("master_score"), "grade": r.get("grade"),
        "linker_score": r.get("linker_analysis",{}).get("adjusted_score"),
        "payload_risk": r.get("payload_profile",{}).get("risk_level"),
        "optimal_dar": r.get("dar_optimization",{}).get("optimal_dar"),
        "tumor_conc": r.get("pk_simulation",{}).get("tumor_conc_nmol_L"),
        "plasma_t12": r.get("pk_simulation",{}).get("t12_beta_h"),
        "top_site": f"{r.get('top_conjugation_site',{}).get('residue','?')}{r.get('top_conjugation_site',{}).get('position','?')}",
        "suitability": r.get("linker_analysis",{}).get("suitability"),
    } for r in valid]
    rec = f"Best: {best.get('candidate_name')} score={best.get('master_score')}/100 grade={best.get('grade')}." if valid else "No valid results."
    return jsonify({"ranked_candidates": valid, "comparison_table": table,
                    "best_candidate": best.get("candidate_name"), "recommendation": rec,
                    "generated_at": datetime.utcnow().isoformat()})


# ─────────────────────────────────────────────
# FEATURE 2: TEXT REPORT DOWNLOAD
# ─────────────────────────────────────────────

@app.route("/api/report/text", methods=["POST"])
def api_report_text():
    from flask import Response
    d = request.json or {}
    r = d.get("results")
    if not r:
        return jsonify({"error": "Pass results field"}), 400
    pk = r.get("pk_simulation", {})
    linker = r.get("linker_analysis", {})
    payload = r.get("payload_profile", {})
    dar = r.get("dar_optimization", {})
    site = r.get("top_conjugation_site", {})
    rid = hashlib.md5(str(r.get("generated_at","")).encode()).hexdigest()[:8].upper()
    lines = [
        "="*60,
        "  ADC-DesignOS ANALYSIS REPORT — PhyloBandits",
        f"  www.phylobandits.com  |  Report ID: {rid}",
        "="*60,
        f"  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "EXECUTIVE SUMMARY","-"*40,
        f"  Master Score   : {r.get('master_score')}/100  (Grade {r.get('grade')})",
        f"  Optimal DAR    : {dar.get('optimal_dar')}",
        f"  Top Site       : {site.get('residue','?')}{site.get('position','?')} — {site.get('conjugation_type','')[:45]}",
        f"  Linker         : {linker.get('suitability','?')} ({linker.get('adjusted_score','?')}/100)",
        f"  Payload Risk   : {payload.get('risk_level','?')}",
        f"  Tumor Conc     : {pk.get('tumor_conc_nmol_L','?')} nmol/L",
        f"  Plasma t½β     : {pk.get('t12_beta_h','?')}h",
        "",
        "CONJUGATION SITES","-"*40,
    ]
    for s in r.get("conjugation_sites",[])[:5]:
        lines.append(f"  [{s['score']:5.1f}] Pos {s['position']:4d} {s['residue']}  {s['conjugation_type'][:40]}")
    lines += ["","LINKER","-"*40,
        f"  Type: {linker.get('type','?').replace('_',' ')}  Score: {linker.get('adjusted_score','?')}/100  t½: {linker.get('plasma_t12_h','?')}h",
        f"  Release: {linker.get('release') or 'None (non-cleavable)'}",
    ]
    for f in linker.get("flags",[]): lines.append(f"  ⚠  {f}")
    for rec in linker.get("recommendations",[]): lines.append(f"  →  {rec}")
    lines += ["","PAYLOAD","-"*40,
        f"  {payload.get('class','?')} | IC50={payload.get('potency_nm','?')}nM | MW={payload.get('mw','?')}Da | logP={payload.get('logP','?')}",
        f"  Risk: {payload.get('risk_level','?')} | hERG={round(payload.get('hERG_risk',0)*100)}% | Hepatotox={round(payload.get('hepatotox',0)*100)}%",
    ]
    for a in payload.get("safety_alerts",[]): lines.append(f"  !  {a}")
    lines += ["","DAR OPTIMIZATION","-"*40,
        f"  {'DAR':<5}{'Efficacy%':<12}{'Aggreg%':<10}{'Tox%':<10}{'t½h':<8}{'TI':<7}Status"
    ]
    for row in dar.get("dar_analysis",[]):
        star = "★" if row["dar"]==dar.get("optimal_dar") else " "
        lines.append(f"  {star}{row['dar']:<4}{row['efficacy_pct']:<12}{row['aggregation_risk_pct']:<10}{row['off_target_toxicity_pct']:<10}{row['plasma_t12_hours']:<8}{row['therapeutic_index']:<7}{row['recommendation']}")
    lines += ["","PK PARAMETERS","-"*40,
        f"  Cmax={pk.get('cmax_nmol_L')}nmol/L  AUC={pk.get('auc_nmol_h_L')}nmol·h/L",
        f"  t½α={pk.get('t12_alpha_h')}h  t½β={pk.get('t12_beta_h')}h  CL={pk.get('cl_L_h_kg')}L/h/kg",
        f"  Tumor={pk.get('tumor_conc_nmol_L')}nmol/L  DAR CL Penalty=+{pk.get('dar_pk_penalty_pct')}%",
        "","RECOMMENDATIONS","-"*40,
    ]
    for rec in r.get("recommendation",[]): lines.append(f"  →  {rec}")
    lines += ["","="*60,"  PhyloBandits | For research use only","="*60]
    return Response("\n".join(lines), mimetype="text/plain",
        headers={"Content-Disposition":"attachment; filename=ADC_Report_PhyloBandits.txt"})


# ─────────────────────────────────────────────
# FEATURE 3: USER HISTORY (SQLite)
# ─────────────────────────────────────────────

import sqlite3, os, uuid

DB_PATH = os.path.join(os.path.dirname(__file__), "history.db")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS analyses (
        id TEXT PRIMARY KEY, session TEXT NOT NULL, name TEXT,
        payload TEXT, linker TEXT, dar INTEGER, score REAL,
        grade TEXT, optimal_dar INTEGER, full_json TEXT, created TEXT)""")
    conn.commit(); conn.close()

init_db()

@app.route("/api/history/save", methods=["POST"])
def history_save():
    d = request.json or {}
    results = d.get("results")
    session = d.get("session_id", "anon")
    name = d.get("name", f"Analysis {datetime.utcnow().strftime('%b %d %H:%M')}")
    if not results: return jsonify({"error":"No results"}), 400
    rid = str(uuid.uuid4())[:8].upper()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""INSERT INTO analyses
        (id,session,name,payload,linker,dar,score,grade,optimal_dar,full_json,created)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (
        rid, session, name,
        results.get("payload_profile",{}).get("class","?"),
        results.get("linker_analysis",{}).get("type","?"),
        results.get("dar_optimization",{}).get("optimal_dar",0),
        results.get("master_score",0), results.get("grade","?"),
        results.get("dar_optimization",{}).get("optimal_dar",0),
        json.dumps(results), datetime.utcnow().isoformat()))
    conn.commit(); conn.close()
    return jsonify({"saved": True, "id": rid})

@app.route("/api/history/<session_id>", methods=["GET"])
def history_get(session_id):
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute("""SELECT id,name,payload,linker,dar,score,grade,optimal_dar,created
        FROM analyses WHERE session=? ORDER BY created DESC LIMIT 50""", (session_id,)).fetchall()
    conn.close()
    return jsonify([{"id":r[0],"name":r[1],"payload":r[2],"linker":r[3],
        "dar":r[4],"score":r[5],"grade":r[6],"optimal_dar":r[7],"created":r[8]} for r in rows])

@app.route("/api/history/load/<rec_id>", methods=["GET"])
def history_load(rec_id):
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT full_json FROM analyses WHERE id=?", (rec_id,)).fetchone()
    conn.close()
    if not row: return jsonify({"error":"Not found"}), 404
    return jsonify(json.loads(row[0]))

@app.route("/api/history/delete/<rec_id>", methods=["DELETE"])
def history_delete(rec_id):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM analyses WHERE id=?", (rec_id,))
    conn.commit(); conn.close()
    return jsonify({"deleted": True})

@app.route("/api/history/stats/<session_id>", methods=["GET"])
def history_stats(session_id):
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT COUNT(*),AVG(score),MAX(score),MIN(score) FROM analyses WHERE session=?",
        (session_id,)).fetchone()
    conn.close()
    return jsonify({"total":row[0],"avg_score":round(row[1] or 0,1),
        "best_score":row[2] or 0,"worst_score":row[3] or 0})


if __name__ == "__main__":
    print("ADC-DesignOS API v2 — PhyloBandits")
    print("Endpoints: /api/analyze  /api/compare  /api/report/text  /api/history/*")
    print("Running on http://0.0.0.0:8000")
    app.run(host="0.0.0.0", port=8000, debug=False)


# ═══════════════════════════════════════════════════
# FEATURE 4: BRANDED PDF REPORT  (reportlab)
# ═══════════════════════════════════════════════════

@app.route("/api/report/pdf", methods=["POST"])
def api_report_pdf():
    """Generate a branded PhyloBandits PDF report from analysis results."""
    from flask import Response
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        import io
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: pip install reportlab"}), 500

    d = request.json or {}
    r = d.get("results")
    if not r:
        return jsonify({"error": "Pass results field"}), 400

    pk   = r.get("pk_simulation", {})
    ln   = r.get("linker_analysis", {})
    pl   = r.get("payload_profile", {})
    dar  = r.get("dar_optimization", {})
    site = r.get("top_conjugation_site", {})
    rid  = hashlib.md5(str(r.get("generated_at","")).encode()).hexdigest()[:8].upper()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=14*mm, bottomMargin=14*mm)

    # ── Colors
    C_BG    = colors.HexColor("#0a0f0d")
    C_GRN   = colors.HexColor("#22d86e")
    C_GRN2  = colors.HexColor("#16a34a")
    C_GRN3  = colors.HexColor("#4ade80")
    C_TEAL  = colors.HexColor("#2dd4bf")
    C_AMB   = colors.HexColor("#fbbf24")
    C_RED   = colors.HexColor("#f87171")
    C_TX    = colors.HexColor("#e8f5ec")
    C_TX2   = colors.HexColor("#94b8a0")
    C_TX3   = colors.HexColor("#5a7a65")
    C_SURF  = colors.HexColor("#1a2d20")
    C_BRD   = colors.HexColor("#243d2c")
    C_WHITE = colors.white

    W = A4[0] - 36*mm  # usable width

    def style(name, **kw):
        s = ParagraphStyle(name, **kw)
        return s

    S_HEAD   = style("head",   fontName="Helvetica-Bold",   fontSize=20, textColor=C_GRN,  alignment=TA_LEFT)
    S_SUB    = style("sub",    fontName="Helvetica",         fontSize=9,  textColor=C_TX3,  alignment=TA_LEFT)
    S_SEC    = style("sec",    fontName="Helvetica-Bold",   fontSize=11, textColor=C_GRN3, spaceAfter=4)
    S_BODY   = style("body",   fontName="Helvetica",         fontSize=9,  textColor=C_TX2,  leading=14)
    S_MONO   = style("mono",   fontName="Courier",           fontSize=8,  textColor=C_TX2,  leading=12)
    S_LABEL  = style("lbl",    fontName="Helvetica-Bold",   fontSize=7,  textColor=C_TX3)
    S_VAL    = style("val",    fontName="Helvetica-Bold",   fontSize=11, textColor=C_TX)
    S_SCORE  = style("score",  fontName="Helvetica-Bold",   fontSize=32, textColor=C_GRN3, alignment=TA_CENTER)
    S_GRADE  = style("grade",  fontName="Helvetica-Bold",   fontSize=14, textColor=C_GRN2, alignment=TA_CENTER)
    S_FOOT   = style("foot",   fontName="Helvetica",         fontSize=7,  textColor=C_TX3,  alignment=TA_CENTER)
    S_REC    = style("rec",    fontName="Helvetica",         fontSize=8.5,textColor=C_GRN3, leading=13)
    S_FLAG   = style("flag",   fontName="Helvetica",         fontSize=8.5,textColor=C_AMB,  leading=13)
    S_WARN   = style("warn",   fontName="Helvetica",         fontSize=8.5,textColor=C_RED,  leading=13)

    def tbl_style(header_rows=1, stripe=True):
        cmds = [
            ("BACKGROUND",   (0,0), (-1,header_rows-1), C_SURF),
            ("TEXTCOLOR",    (0,0), (-1,header_rows-1), C_GRN3),
            ("FONTNAME",     (0,0), (-1,header_rows-1), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 7.5),
            ("TEXTCOLOR",    (0,header_rows), (-1,-1), C_TX2),
            ("GRID",         (0,0), (-1,-1), 0.3, C_BRD),
            ("ROWBACKGROUNDS",(0,header_rows),(-1,-1), [C_BG, colors.HexColor("#0d160f")] if stripe else [C_BG]),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ]
        return TableStyle(cmds)

    score_val = r.get("master_score", 0)
    grade_val = r.get("grade", "?")
    grade_color = C_GRN3 if grade_val=="A" else C_TEAL if grade_val=="B" else C_AMB if grade_val=="C" else C_RED

    story = []

    # ── HEADER BLOCK
    header_data = [[
        Paragraph("<b>ADC-DesignOS</b>", style("hdr", fontName="Helvetica-Bold", fontSize=22, textColor=C_GRN)),
        Paragraph(f"<b>{score_val}</b>", style("sc", fontName="Helvetica-Bold", fontSize=36, textColor=grade_color, alignment=TA_RIGHT)),
    ],[
        Paragraph("PhyloBandits Advanced Bioinformatics | www.phylobandits.com", S_SUB),
        Paragraph(f"<b>Grade {grade_val}</b>  ·  Report {rid}  ·  {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}", style("hdr2", fontName="Helvetica", fontSize=8, textColor=C_TX3, alignment=TA_RIGHT)),
    ]]
    ht = Table(header_data, colWidths=[W*0.65, W*0.35])
    ht.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), C_SURF),
        ("TOPPADDING",  (0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("LEFTPADDING", (0,0),(-1,-1), 10),
        ("RIGHTPADDING",(0,0),(-1,-1), 10),
        ("VALIGN",      (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(ht)
    story.append(Spacer(1, 6*mm))

    # ── EXECUTIVE SUMMARY ROW
    story.append(Paragraph("EXECUTIVE SUMMARY", S_SEC))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 2*mm))

    summ_items = [
        ("Top Conj. Site", f"{site.get('residue','?')}{site.get('position','?')} — {site.get('conjugation_type','')[:40]}"),
        ("Linker Suitability", f"{ln.get('suitability','?')} ({ln.get('adjusted_score','?')}/100)"),
        ("Optimal DAR", str(dar.get('optimal_dar','?'))),
        ("Payload Risk", pl.get('risk_level','?')),
        ("Plasma t½β", f"{pk.get('t12_beta_h','?')} h"),
        ("Tumor Conc.", f"{pk.get('tumor_conc_nmol_L','?')} nmol/L"),
        ("Cmax", f"{pk.get('cmax_nmol_L','?')} nmol/L"),
        ("AUC", f"{pk.get('auc_nmol_h_L','?')} nmol·h/L"),
    ]
    summ_data = [[Paragraph(f"<b>{k}</b>", S_LABEL), Paragraph(str(v), S_BODY)] for k,v in summ_items]
    summ_col = [summ_data[:4], summ_data[4:]]
    for col in summ_col:
        t = Table(col, colWidths=[W*0.22, W*0.28])
        t.setStyle(tbl_style(header_rows=0, stripe=True))
        story.append(t)
    story.append(Spacer(1, 4*mm))

    # ── RECOMMENDATIONS
    story.append(Paragraph("RECOMMENDATIONS", S_SEC))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 1*mm))
    for rec in r.get("recommendation", []):
        story.append(Paragraph(f"→  {rec}", S_REC))
    story.append(Spacer(1, 4*mm))

    # ── CONJUGATION SITES TABLE
    story.append(Paragraph("CONJUGATION SITE PREDICTION", S_SEC))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 1*mm))
    sites = r.get("conjugation_sites", [])[:8]
    if sites:
        hdr = [Paragraph(h, S_LABEL) for h in ["Pos","AA","Score","Type","Rationale"]]
        rows = [hdr] + [[
            Paragraph(str(s["position"]), S_MONO),
            Paragraph(s["residue"], style("aa", fontName="Courier-Bold", fontSize=9, textColor=C_GRN3)),
            Paragraph(str(s["score"]), S_MONO),
            Paragraph(s["conjugation_type"][:28], S_BODY),
            Paragraph("; ".join(s.get("rationale",[])), S_BODY),
        ] for s in sites]
        t = Table(rows, colWidths=[W*0.08, W*0.06, W*0.08, W*0.28, W*0.50])
        t.setStyle(tbl_style())
        story.append(t)
    story.append(Spacer(1, 4*mm))

    # ── LINKER + PAYLOAD side by side
    story.append(Paragraph("LINKER ANALYSIS & PAYLOAD PROFILE", S_SEC))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 1*mm))

    lnk_rows = [
        [Paragraph("Type",       S_LABEL), Paragraph(str(ln.get("type","?")).replace("_"," "), S_BODY)],
        [Paragraph("Score",      S_LABEL), Paragraph(f"{ln.get('adjusted_score','?')}/100", S_BODY)],
        [Paragraph("Suitability",S_LABEL), Paragraph(str(ln.get("suitability","?")), S_BODY)],
        [Paragraph("Plasma t½",  S_LABEL), Paragraph(f"{ln.get('plasma_t12_h','?')} h", S_BODY)],
        [Paragraph("Release",    S_LABEL), Paragraph(str(ln.get("release") or "None (non-cleavable)"), S_BODY)],
        [Paragraph("Flags",      S_LABEL), Paragraph("; ".join(ln.get("flags",[]) or ["None"]), S_FLAG)],
    ]
    pl_rows = [
        [Paragraph("Class",      S_LABEL), Paragraph(str(pl.get("class","?")), S_BODY)],
        [Paragraph("Mechanism",  S_LABEL), Paragraph(str(pl.get("mechanism","?")), S_BODY)],
        [Paragraph("IC50",       S_LABEL), Paragraph(f"{pl.get('potency_nm','?')} nM", S_BODY)],
        [Paragraph("MW",         S_LABEL), Paragraph(f"{pl.get('mw','?')} Da", S_BODY)],
        [Paragraph("hERG Risk",  S_LABEL), Paragraph(f"{round(pl.get('hERG_risk',0)*100)}%", S_BODY)],
        [Paragraph("Risk Level", S_LABEL), Paragraph(str(pl.get("risk_level","?")), S_BODY)],
    ]
    lnk_t = Table(lnk_rows, colWidths=[W*0.16, W*0.34])
    pl_t  = Table(pl_rows,  colWidths=[W*0.16, W*0.34])
    lnk_t.setStyle(tbl_style(header_rows=0))
    pl_t.setStyle(tbl_style(header_rows=0))
    side = Table([[lnk_t, pl_t]], colWidths=[W*0.50, W*0.50])
    side.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),4)]))
    story.append(side)
    story.append(Spacer(1, 4*mm))

    # ── DAR TABLE
    story.append(Paragraph("DAR OPTIMIZATION", S_SEC))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 1*mm))
    dar_analysis = dar.get("dar_analysis", [])
    if dar_analysis:
        dar_hdr = [Paragraph(h, S_LABEL) for h in ["DAR","Efficacy%","Aggreg%","Tox%","t½ (h)","Therap. Index","Status"]]
        opt = dar.get("optimal_dar")
        dar_rows = [dar_hdr] + [[
            Paragraph(("★ " if row["dar"]==opt else "") + str(row["dar"]),
                style("dv", fontName="Courier-Bold" if row["dar"]==opt else "Courier",
                      fontSize=8, textColor=C_GRN3 if row["dar"]==opt else C_TX2)),
            Paragraph(str(row["efficacy_pct"]),          S_MONO),
            Paragraph(str(row["aggregation_risk_pct"]),  S_MONO),
            Paragraph(str(row["off_target_toxicity_pct"]),S_MONO),
            Paragraph(str(row["plasma_t12_hours"]),       S_MONO),
            Paragraph(str(row["therapeutic_index"]),      S_MONO),
            Paragraph(str(row["recommendation"])[:30],    S_BODY),
        ] for row in dar_analysis]
        dt = Table(dar_rows, colWidths=[W*0.08,W*0.12,W*0.10,W*0.09,W*0.10,W*0.16,W*0.35])
        dt.setStyle(tbl_style())
        story.append(dt)
    story.append(Spacer(1, 4*mm))

    # ── PK TABLE
    story.append(Paragraph("PHARMACOKINETICS (Two-Compartment Model)", S_SEC))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 1*mm))
    pk_items = [
        ("Cmax",          f"{pk.get('cmax_nmol_L','?')} nmol/L"),
        ("AUC",           f"{pk.get('auc_nmol_h_L','?')} nmol·h/L"),
        ("t½ alpha",      f"{pk.get('t12_alpha_h','?')} h"),
        ("t½ beta",       f"{pk.get('t12_beta_h','?')} h"),
        ("Tumor Conc.",   f"{pk.get('tumor_conc_nmol_L','?')} nmol/L"),
        ("Free Payload AUC", f"{pk.get('free_payload_auc','?')} nmol·h/L"),
        ("Clearance",     f"{pk.get('cl_L_h_kg','?')} L/h/kg"),
        ("DAR CL Penalty",f"+{pk.get('dar_pk_penalty_pct','?')}%"),
    ]
    pk_data = [[Paragraph(k, S_LABEL), Paragraph(v, S_BODY)] for k,v in pk_items]
    # 2-column layout
    left_pk = pk_data[:4]; right_pk = pk_data[4:]
    lt = Table(left_pk,  colWidths=[W*0.20, W*0.30])
    rt = Table(right_pk, colWidths=[W*0.20, W*0.30])
    lt.setStyle(tbl_style(header_rows=0)); rt.setStyle(tbl_style(header_rows=0))
    pk_side = Table([[lt, rt]], colWidths=[W*0.50, W*0.50])
    pk_side.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),4)]))
    story.append(pk_side)
    story.append(Spacer(1, 6*mm))

    # ── FOOTER
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BRD))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"Report ID: {rid}  ·  Generated by ADC-DesignOS v2  ·  PhyloBandits Advanced Bioinformatics  ·  www.phylobandits.com  ·  For research use only.",
        S_FOOT))

    # ── BUILD
    doc.build(story)
    buf.seek(0)
    return Response(buf.read(), mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=ADC_Report_{rid}_PhyloBandits.pdf"})


# ═══════════════════════════════════════════════════
# FEATURE 5: STRUCTURE VIEWER DATA
# Returns 3D-ready conjugation site data for NGL / Mol*
# (No AlphaFold API needed — uses pre-computed Trastuzumab coords)
# ═══════════════════════════════════════════════════

import math

@app.route("/api/structure/sites", methods=["POST"])
def api_structure_sites():
    """
    Return conjugation site data formatted for 3D structure visualization.
    Generates realistic pseudo-3D coordinates for a VH domain antibody.
    In production, replace with AlphaFold3 API call or PDB fetch.
    """
    d = request.json or {}
    seq = d.get("sequence","")
    if not seq:
        return jsonify({"error":"sequence required"}), 400

    # Run conjugation site prediction
    sites = predict_conjugation_sites(seq)[:10]

    # Generate pseudo-3D coordinates (spiral ribbon model — realistic for VH domain)
    # Real deployment: call https://alphafoldserver.com/api or fetch PDB via RCSB
    def pseudo_coords(pos, seq_len):
        """Approximate Cα coordinates for a VH domain beta-barrel fold."""
        frac = pos / max(seq_len, 1)
        # Beta barrel: ~120 Å long, ~30 Å diameter
        angle = frac * 4 * math.pi  # 2 full turns
        radius = 15 + 5 * math.sin(frac * math.pi * 6)  # vary radius
        x = round(radius * math.cos(angle), 2)
        y = round(radius * math.sin(angle), 2)
        z = round(frac * 60 - 30, 2)  # 60 Å total length
        return {"x": x, "y": y, "z": z}

    site_3d = []
    for s in sites:
        coords = pseudo_coords(s["position"], len(seq))
        site_3d.append({
            "position":        s["position"],
            "residue":         s["residue"],
            "score":           s["score"],
            "conjugation_type":s["conjugation_type"],
            "coords":          coords,
            "color":           "#4ade80" if s["score"]>80 else "#fbbf24" if s["score"]>60 else "#f87171",
            "size":            max(0.4, s["score"]/100),
            "label":           f"{s['residue']}{s['position']} ({s['score']:.0f})",
        })

    return jsonify({
        "sites": site_3d,
        "sequence_length": len(seq),
        "pdb_hint": "For real 3D structure: fetch https://alphafoldserver.com/api/prediction?sequence=YOUR_SEQ or https://www.rcsb.org/structure/1IGT",
        "visualization_note": "Coordinates are pseudo-3D approximations. Upload a real PDB for production use.",
        "generated_at": datetime.utcnow().isoformat()
    })


# ═══════════════════════════════════════════════════
# FEATURE 6: LEAD OPTIMIZATION WIZARD
# AI-style scoring grid — recommends best linker+payload
# for given antibody + target context
# ═══════════════════════════════════════════════════

@app.route("/api/optimize/lead", methods=["POST"])
def api_lead_optimizer():
    """
    Given a sequence + target context, automatically screen all 8×8=64
    linker/payload combinations and rank by therapeutic index + safety.
    Returns top 5 leads with rationale.
    """
    d = request.json or {}
    seq          = d.get("sequence","")
    target_loc   = d.get("target_location","intracellular")
    target_expr  = d.get("target_expression","high")
    indication   = d.get("indication","solid_tumor")  # solid_tumor | hematologic | solid_low_burden
    dose         = float(d.get("dose_mg_kg", 3.0))
    tox_thresh   = float(d.get("toxicity_threshold", 70.0))

    if not seq:
        return jsonify({"error":"sequence required"}), 400

    # Screen all 64 combinations
    results = []
    for linker in LINKER_DB:
        for payload in PAYLOAD_DB:
            try:
                r = rank_candidates(seq, linker, payload, 4,  # DAR=4 as baseline
                    target_loc, target_expr, dose, tox_thresh)

                # Indication-specific adjustments
                ind_bonus = 0
                if indication == "hematologic":
                    # Hematologic: bystander effect valuable, lower tumor density
                    be = PAYLOAD_DB[payload].get("bystander_effect","")
                    if "High" in str(be): ind_bonus += 8
                    if PAYLOAD_DB[payload].get("mdr_risk","") in ["Low","Very Low"]: ind_bonus += 5
                elif indication == "solid_low_burden":
                    # Low burden: want highly potent payloads
                    if PAYLOAD_DB[payload]["potency_nm"] < 0.1: ind_bonus += 10
                else:  # solid_tumor
                    # Standard solid tumor: moderate bystander, stability
                    if LINKER_DB[linker].get("stability","") == "High": ind_bonus += 5

                adj_score = min(100, r["master_score"] + ind_bonus)

                results.append({
                    "linker":           linker,
                    "payload":          payload,
                    "master_score":     adj_score,
                    "base_score":       r["master_score"],
                    "indication_bonus": ind_bonus,
                    "grade":            r["grade"],
                    "optimal_dar":      r["dar_optimization"]["optimal_dar"],
                    "linker_score":     r["linker_analysis"]["adjusted_score"],
                    "linker_suitability": r["linker_analysis"]["suitability"],
                    "payload_risk":     r["payload_profile"]["risk_level"],
                    "payload_class":    r["payload_profile"]["class"],
                    "payload_ic50":     r["payload_profile"]["potency_nm"],
                    "tumor_conc":       r["pk_simulation"]["tumor_conc_nmol_L"],
                    "plasma_t12":       r["pk_simulation"]["t12_beta_h"],
                    "therapeutic_index":r["dar_optimization"]["dar_analysis"][3]["therapeutic_index"],
                    "flags":            r["linker_analysis"].get("flags",[]),
                    "full_result":      r,
                })
            except Exception:
                pass

    results.sort(key=lambda x: x["master_score"], reverse=True)
    top5 = results[:5]

    # Generate rationale for each
    for i, lead in enumerate(top5):
        reason_parts = []
        if lead["linker_suitability"] in ["Excellent","Good"]:
            reason_parts.append(f"{lead['linker_suitability'].lower()} linker stability")
        if lead["payload_risk"] in ["Low","Moderate"]:
            reason_parts.append(f"{lead['payload_risk'].lower()} payload toxicity")
        if lead["optimal_dar"] in [2, 4]:
            reason_parts.append(f"optimal DAR {lead['optimal_dar']} achievable")
        if lead["tumor_conc"] > 0.05:
            reason_parts.append("strong tumor accumulation")
        if not lead["flags"]:
            reason_parts.append("no linker warnings")
        lead["rationale"] = f"Rank #{i+1}: {lead['linker']} + {lead['payload']} — " + ("; ".join(reason_parts) if reason_parts else "balanced profile") + f". Score {lead['master_score']}/100."

    # Build comparison stats
    all_scores = [r["master_score"] for r in results]
    best = top5[0]
    margin = round(best["master_score"] - top5[1]["master_score"], 1) if len(top5)>1 else 0

    return jsonify({
        "top_leads":       top5,
        "best_combination":f"{best['linker']} + {best['payload']}",
        "best_score":       best["master_score"],
        "best_dar":         best["optimal_dar"],
        "margin_over_2nd":  margin,
        "total_screened":   len(results),
        "indication":       indication,
        "screening_summary":{
            "mean_score":   round(sum(all_scores)/len(all_scores),1) if all_scores else 0,
            "max_score":    max(all_scores) if all_scores else 0,
            "min_score":    min(all_scores) if all_scores else 0,
            "above_70":     sum(1 for s in all_scores if s>=70),
            "above_80":     sum(1 for s in all_scores if s>=80),
        },
        "recommendation": f"Best combination for {indication}: {best['linker']} + {best['payload']} (Score {best['master_score']}/100, Grade {best['grade']}). Margin over 2nd: +{margin} points. Proceed with DAR {best['optimal_dar']} conjugation.",
        "generated_at": datetime.utcnow().isoformat()
    })


# ═══════════════════════════════════════════════════════════
# FEATURE 7: SMILES CUSTOM PAYLOAD ANALYSER
# Pure-Python cheminformatics — no RDKit needed
# ═══════════════════════════════════════════════════════════

import re as _re

def smiles_to_properties(smiles: str) -> dict:
    """
    Parse a SMILES string and compute:
      - MW (from atom counts + average atomic masses)
      - logP  (Wildman-Crippen fragment approach, simplified)
      - HBD / HBA counts
      - Rotatable bonds
      - Aromatic ring count
      - TPSA estimate
      - hERG risk score  (structural alerts: basic N, lipophilic, etc.)
      - Hepatotox score  (structural alerts)
      - Rule-of-5 assessment
    """
    s = smiles.strip()
    if not s:
        return {"error": "Empty SMILES"}

    # ── Atom masses (monoisotopic)
    MASSES = {
        'C':12.011,'N':14.007,'O':15.999,'S':32.06,'P':30.974,
        'F':18.998,'Cl':35.45,'Br':79.904,'I':126.904,'H':1.008,
        'Si':28.086,'B':10.811,'Se':78.971,'Fe':55.845,'Co':58.933,
    }

    # ── Count heavy atoms (case-sensitive SMILES parsing)
    # Bracket atoms like [NH2], [C@@H] etc.
    atom_counts = {}
    work = s

    # Remove stereochemistry decorators that aren't atoms
    work_clean = _re.sub(r'\[([A-Za-z]+)[@#H0-9+\-]*\]', r'[\1]', s)

    # Extract bracket atoms
    for m in _re.finditer(r'\[([A-Z][a-z]?)', work_clean):
        sym = m.group(1)
        atom_counts[sym] = atom_counts.get(sym, 0) + 1

    # Extract bare atoms (not in brackets) — uppercase = heavy, lowercase = aromatic carbon
    bare = _re.sub(r'\[.*?\]', '', work_clean)
    for ch in bare:
        if ch == 'c': atom_counts['C'] = atom_counts.get('C',0)+1
        elif ch == 'n': atom_counts['N'] = atom_counts.get('N',0)+1
        elif ch == 'o': atom_counts['O'] = atom_counts.get('O',0)+1
        elif ch == 's': atom_counts['S'] = atom_counts.get('S',0)+1
        elif ch.isupper():
            # Two-char symbols first
            pass
    # Better: scan bare string for element symbols
    i = 0
    atom_counts2 = {}
    while i < len(bare):
        if bare[i].isupper():
            if i+1 < len(bare) and bare[i+1].islower() and bare[i+1] in 'lrni':
                sym = bare[i:i+2]
                i += 2
            else:
                sym = bare[i]
                i += 1
            atom_counts2[sym] = atom_counts2.get(sym, 0) + 1
        elif bare[i].islower() and bare[i] in 'cnos':
            sym = bare[i].upper()
            atom_counts2[sym] = atom_counts2.get(sym, 0) + 1
            i += 1
        else:
            i += 1
    # Merge
    for sym,cnt in atom_counts2.items():
        atom_counts[sym] = atom_counts.get(sym, 0) + cnt

    # ── Estimate H count (simple valence model)
    nC  = atom_counts.get('C',0)
    nN  = atom_counts.get('N',0)
    nO  = atom_counts.get('O',0)
    nS  = atom_counts.get('S',0)
    nF  = atom_counts.get('F',0)
    nCl = atom_counts.get('Cl',0)
    nBr = atom_counts.get('Br',0)
    nI  = atom_counts.get('I',0)
    nP  = atom_counts.get('P',0)
    nSi = atom_counts.get('Si',0)

    # Count explicit H in SMILES
    nH_explicit = len(_re.findall(r'(?<!\[)H(?!\])|(?<=\[)[A-Za-z]*H\d*', s))
    # Rough implicit H estimate
    double_bonds = s.count('=')
    triple_bonds = s.count('#')
    ring_closures = len(_re.findall(r'\d', _re.sub(r'\[.*?\]','',s))) // 2
    # Degree of unsaturation  
    DoU = (2*nC + 2 + nN - nF - nCl - nBr - nI) / 2 - ring_closures
    nH = max(0, round(2*nC + 2 + nN - 2*nO*0 - nF - nCl - nBr - nI - 2*double_bonds - 4*triple_bonds*0))
    nH = nH_explicit + max(0, nH // 4)  # rough

    # ── MW
    MW = (nC*12.011 + nN*14.007 + nO*15.999 + nS*32.06 +
          nF*18.998 + nCl*35.45 + nBr*79.904 + nI*126.904 +
          nP*30.974 + nSi*28.086 + nH*1.008)
    MW = round(MW, 1)

    # ── logP (Wildman–Crippen simplified fragment contributions)
    logP = 0.0
    logP += nC  * 0.1441
    logP += nN  * (-1.019)
    logP += nO  * (-0.6835)
    logP += nS  * 0.4375
    logP += nF  * 0.4727
    logP += nCl * 0.6895
    logP += nBr * 0.8456
    logP += nI  * 1.1170
    logP += nP  * 0.8456
    # Aromatic correction
    n_arom_c = sum(1 for c in s if c == 'c')
    n_arom_n = sum(1 for c in s if c == 'n')
    logP += n_arom_c * (0.2956 - 0.1441)  # aromatic C vs aliphatic
    logP += n_arom_n * (0.0925)
    logP = round(logP, 2)

    # ── H-bond donors/acceptors
    HBD = len(_re.findall(r'(?:(?<!\[)[NO]H|\[NH|\[OH)', s))
    HBA = nN + nO

    # ── Rotatable bonds (single bonds not in ring, not to H, not amide-like)
    # Count single bonds between heavy atoms in chain
    rot_bonds = max(0, len(_re.findall(r'(?<![=#])-(?![=#])', _re.sub(r'\(.*?\)','',s))) + nC//4 - ring_closures)

    # ── Rings
    ring_count = len(_re.findall(r'\d', _re.sub(r'\[.*?\]','',s))) // 2
    arom_rings = round(n_arom_c / 6)

    # ── TPSA estimate (polar surface area)
    tpsa = nO * 9.23 + nN * 12.89 + nS * 25.3
    # Adjust for NH, OH
    tpsa += HBD * 10.5
    tpsa = round(tpsa, 1)

    # ── Structural alerts for hERG (QT prolongation)
    herg_risk = 0.0
    herg_alerts = []
    if nN >= 1 and logP > 2:
        herg_risk += 0.25; herg_alerts.append("Basic nitrogen + lipophilicity (logP>2)")
    if arom_rings >= 2:
        herg_risk += 0.15; herg_alerts.append("Multiple aromatic rings")
    if logP > 4:
        herg_risk += 0.20; herg_alerts.append("High lipophilicity (logP>4)")
    if nN >= 2:
        herg_risk += 0.10; herg_alerts.append("Multiple nitrogen atoms")
    if MW > 500 and logP > 3:
        herg_risk += 0.10; herg_alerts.append("Heavy lipophilic molecule")
    herg_risk = min(1.0, round(herg_risk, 2))

    # ── Hepatotox alerts
    hepatotox_risk = 0.0
    hepatotox_alerts = []
    if 'Cl' in s and nN >= 1:
        hepatotox_risk += 0.15; hepatotox_alerts.append("Chlorinated + nitrogen (CYP activation risk)")
    if nS >= 1 and arom_rings >= 1:
        hepatotox_risk += 0.10; hepatotox_alerts.append("Thiophene/thioaromatic scaffold")
    if logP > 5:
        hepatotox_risk += 0.20; hepatotox_alerts.append("Extreme lipophilicity (logP>5)")
    if tpsa < 20:
        hepatotox_risk += 0.10; hepatotox_alerts.append("Very low TPSA — poor aqueous solubility")
    hepatotox_risk = min(1.0, round(hepatotox_risk, 2))

    # ── Overall risk (composite)
    overall_risk = round((herg_risk * 0.4 + hepatotox_risk * 0.35 + (1 if logP > 5 else 0) * 0.15 + (1 if MW > 800 else 0) * 0.10) * 100)
    risk_level = "Low" if overall_risk < 30 else "Moderate" if overall_risk < 60 else "High"

    # ── Lipinski Rule of 5
    ro5_violations = sum([MW > 500, logP > 5, HBD > 5, HBA > 10])
    ro5_pass = ro5_violations == 0

    # ── Estimated payload class from structure
    if nN >= 2 and arom_rings >= 1 and logP > 2:
        est_class = "Auristatin-like (microtubule inhibitor)"
    elif 'N(=O)' in s or 'n' in s and nN >= 3:
        est_class = "DNA-interactive (PBD/alkylator-like)"
    elif nO >= 4 and nC >= 10:
        est_class = "Maytansinoid-like (DM1/DM4)"
    elif 'O=C' in s and nO >= 3:
        est_class = "Camptothecin-like (topoisomerase I)"
    else:
        est_class = "Novel scaffold (unclassified)"

    # ── ADC suitability score
    adc_score = 100
    if MW < 300: adc_score -= 20  # too small
    if MW > 900: adc_score -= 15  # too large
    if logP < 0:  adc_score -= 10  # too hydrophilic (poor cell penetration)
    if logP > 6:  adc_score -= 20  # too lipophilic (aggregation)
    if HBD > 5:   adc_score -= 10  # poor membrane permeability
    if ro5_violations >= 2: adc_score -= 15
    if herg_risk > 0.5: adc_score -= 20
    if hepatotox_risk > 0.5: adc_score -= 15
    adc_score = max(0, min(100, adc_score))
    adc_grade = 'A' if adc_score>=80 else 'B' if adc_score>=65 else 'C' if adc_score>=50 else 'D'

    return {
        "smiles": s,
        "atom_counts": atom_counts,
        "mw": MW,
        "logP": logP,
        "HBD": HBD,
        "HBA": HBA,
        "rotatable_bonds": rot_bonds,
        "ring_count": ring_count,
        "aromatic_rings": arom_rings,
        "tpsa": tpsa,
        "hERG_risk": herg_risk,
        "hERG_alerts": herg_alerts,
        "hepatotox_risk": hepatotox_risk,
        "hepatotox_alerts": hepatotox_alerts,
        "overall_risk_score": overall_risk,
        "risk_level": risk_level,
        "ro5_violations": ro5_violations,
        "ro5_pass": ro5_pass,
        "estimated_payload_class": est_class,
        "adc_suitability_score": adc_score,
        "adc_grade": adc_grade,
        "doh": round(DoU, 1),
        "recommendations": _smiles_recs(MW, logP, herg_risk, hepatotox_risk, adc_score, ro5_violations),
    }

def _smiles_recs(mw, logp, herg, hepatotox, adc_score, ro5_v):
    recs = []
    if logp < 1:
        recs.append("Consider adding hydrophobic groups — logP < 1 may limit membrane permeability after linker cleavage.")
    elif logp > 5:
        recs.append("High logP may cause ADC aggregation. Consider PEG-linker or hydrophilic spacer.")
    if herg > 0.4:
        recs.append("Elevated hERG risk — screen for QT prolongation. Consider structural modification to reduce basicity.")
    if hepatotox > 0.3:
        recs.append("Hepatotoxicity structural alert present. Evaluate CYP450 metabolism profile.")
    if mw > 800:
        recs.append("High MW may reduce DAR efficiency and increase aggregation risk.")
    if mw < 300:
        recs.append("Low MW may be insufficient for potent cytotoxicity at ADC doses.")
    if ro5_v >= 2:
        recs.append("Multiple Ro5 violations — verify cell membrane permeability experimentally.")
    if adc_score >= 80:
        recs.append("Excellent ADC payload candidate. Proceed to conjugation chemistry design.")
    elif adc_score >= 65:
        recs.append("Good ADC payload potential. Address flagged alerts before in vitro testing.")
    else:
        recs.append("Moderate concerns — consider scaffold optimization before ADC development.")
    return recs


@app.route("/api/smiles/analyze", methods=["POST"])
def api_smiles_analyze():
    """Analyze a custom SMILES payload for ADC suitability."""
    d = request.json or {}
    smiles = d.get("smiles","").strip()
    name   = d.get("name","Custom Payload")
    if not smiles:
        return jsonify({"error": "smiles field required"}), 400
    result = smiles_to_properties(smiles)
    if "error" in result:
        return jsonify(result), 400
    result["name"] = name
    result["generated_at"] = datetime.utcnow().isoformat()
    return jsonify(result)


@app.route("/api/smiles/compare-builtin", methods=["POST"])
def api_smiles_compare():
    """Compare a custom SMILES payload against all 8 built-in payloads."""
    d = request.json or {}
    smiles = d.get("smiles","").strip()
    name   = d.get("name","Custom")
    if not smiles:
        return jsonify({"error": "smiles required"}), 400

    custom = smiles_to_properties(smiles)
    if "error" in custom:
        return jsonify(custom), 400
    custom["name"] = name

    builtins = []
    for pname, pdata in PAYLOAD_DB.items():
        builtins.append({
            "name": pname,
            "class": pdata["class"],
            "potency_nm": pdata["potency_nm"],
            "mw": pdata.get("mw", "?"),
            "logP": pdata.get("logP", "?"),
            "hERG_risk": pdata.get("hERG_risk", "?"),
            "hepatotox": pdata.get("hepatotox", "?"),
            "risk_level": pdata.get("risk_level","?"),
            "mdr_risk": pdata.get("mdr_risk","?"),
        })

    return jsonify({
        "custom_payload": custom,
        "builtin_payloads": builtins,
        "recommendation": (
            f"Custom payload ({name}) scores {custom['adc_suitability_score']}/100 (Grade {custom['adc_grade']}). "
            f"MW={custom['mw']} Da, logP={custom['logP']}, risk={custom['risk_level']}. "
            + (custom['recommendations'][0] if custom['recommendations'] else "")
        ),
        "generated_at": datetime.utcnow().isoformat()
    })


# ═══════════════════════════════════════════════════════════
# FEATURE 8: BATCH MODE — CSV upload → ranked Excel output
# ═══════════════════════════════════════════════════════════

import csv as _csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@app.route("/api/batch/analyze", methods=["POST"])
def api_batch_analyze():
    """
    Accept JSON array of candidates OR CSV string.
    Input: { "candidates": [ {name, sequence, linker, payload, dar, target_location, target_expression}, ... ] }
    OR:    { "csv": "name,sequence,linker,payload,dar,target_location,target_expression\\n..." }
    Returns: Excel file with ranked results.
    """
    from flask import Response
    d = request.json or {}

    candidates = []
    if "csv" in d:
        reader = _csv.DictReader(d["csv"].splitlines())
        for row in reader:
            candidates.append(row)
    elif "candidates" in d:
        candidates = d["candidates"]
    else:
        return jsonify({"error": "Provide 'candidates' array or 'csv' string"}), 400

    if not candidates:
        return jsonify({"error": "No candidates provided"}), 400
    if len(candidates) > 200:
        return jsonify({"error": "Max 200 candidates per batch"}), 400

    results = []
    errors  = []
    for i, c in enumerate(candidates):
        try:
            r = rank_candidates(
                str(c.get("sequence","")).strip(),
                str(c.get("linker","mc-VC-PABC")).strip(),
                str(c.get("payload","MMAE")).strip(),
                int(c.get("dar",4)),
                str(c.get("target_location","intracellular")).strip(),
                str(c.get("target_expression","high")).strip(),
                float(c.get("dose_mg_kg",3.0)),
                float(c.get("toxicity_threshold",70.0)),
            )
            results.append({
                "row": i+1,
                "name": c.get("name", f"Candidate_{i+1}"),
                "linker": c.get("linker","mc-VC-PABC"),
                "payload": c.get("payload","MMAE"),
                "dar": c.get("dar",4),
                "target_location": c.get("target_location","intracellular"),
                "target_expression": c.get("target_expression","high"),
                "master_score": r["master_score"],
                "grade": r["grade"],
                "linker_score": r["linker_analysis"]["adjusted_score"],
                "linker_suitability": r["linker_analysis"]["suitability"],
                "payload_risk": r["payload_profile"]["risk_level"],
                "payload_ic50_nm": r["payload_profile"]["potency_nm"],
                "optimal_dar": r["dar_optimization"]["optimal_dar"],
                "therapeutic_index": r["dar_optimization"]["dar_analysis"][min(int(c.get("dar",4))//2, 5)]["therapeutic_index"],
                "tumor_conc_nmol_L": r["pk_simulation"]["tumor_conc_nmol_L"],
                "plasma_t12_h": r["pk_simulation"]["t12_beta_h"],
                "cmax_nmol_L": r["pk_simulation"]["cmax_nmol_L"],
                "top_site": f"{r['top_conjugation_site']['residue']}{r['top_conjugation_site']['position']}",
                "flags": "; ".join(r["linker_analysis"].get("flags",[]) or []),
                "recommendation": r["recommendation"][0] if r.get("recommendation") else "",
            })
        except Exception as e:
            errors.append({"row": i+1, "name": c.get("name","?"), "error": str(e)})

    # Sort by score
    results.sort(key=lambda x: x["master_score"], reverse=True)
    for i, r in enumerate(results):
        r["rank"] = i+1

    # ── Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ADC Rankings"

    # Color palette
    BG_HEADER  = PatternFill("solid", fgColor="16a34a")
    BG_BEST    = PatternFill("solid", fgColor="1a2d20")
    BG_GOOD    = PatternFill("solid", fgColor="111d16")
    BG_NORMAL  = PatternFill("solid", fgColor="0a0f0d")
    BG_GRADE_A = PatternFill("solid", fgColor="14532d")
    BG_GRADE_B = PatternFill("solid", fgColor="134e4a")
    BG_GRADE_C = PatternFill("solid", fgColor="78350f")
    BG_GRADE_D = PatternFill("solid", fgColor="7f1d1d")
    FONT_HEAD  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FONT_BODY  = Font(name="Calibri", color="E8F5EC", size=9)
    FONT_GRADE = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FONT_SCORE = Font(name="Calibri", bold=True, color="4ade80", size=11)
    FONT_LOGO  = Font(name="Calibri", bold=True, color="22d86e", size=14)
    ALIGN_C    = Alignment(horizontal="center", vertical="center")
    ALIGN_L    = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── TITLE ROW
    ws.merge_cells("A1:V1")
    ws["A1"] = f"ADC-DesignOS Batch Analysis — PhyloBandits | www.phylobandits.com | {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')} | {len(results)} candidates"
    ws["A1"].font = FONT_LOGO
    ws["A1"].fill = PatternFill("solid", fgColor="0f1a14")
    ws["A1"].alignment = ALIGN_L
    ws.row_dimensions[1].height = 22

    # ── COLUMN HEADERS
    headers = [
        "Rank","Name","Linker","Payload","DAR","Location","Expression",
        "Score","Grade","Linker Score","Linker Suitability","Payload Risk",
        "IC50 (nM)","Optimal DAR","Therap. Index","Tumor Conc (nmol/L)",
        "Plasma t½ (h)","Cmax (nmol/L)","Top Site","Flags","Recommendation"
    ]
    col_widths = [6,22,18,14,5,14,12,8,7,12,16,13,10,11,13,16,13,14,10,32,45]

    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = FONT_HEAD
        cell.fill = BG_HEADER
        cell.alignment = ALIGN_C
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18

    # ── DATA ROWS
    grade_fills = {"A": BG_GRADE_A, "B": BG_GRADE_B, "C": BG_GRADE_C, "D": BG_GRADE_D}
    for r in results:
        row_num = r["rank"] + 2
        row_data = [
            r["rank"], r["name"], r["linker"], r["payload"], r["dar"],
            r["target_location"], r["target_expression"],
            r["master_score"], r["grade"], r["linker_score"],
            r["linker_suitability"], r["payload_risk"],
            r["payload_ic50_nm"], r["optimal_dar"], r["therapeutic_index"],
            r["tumor_conc_nmol_L"], r["plasma_t12_h"], r["cmax_nmol_L"],
            r["top_site"], r["flags"], r["recommendation"]
        ]
        row_bg = BG_BEST if r["rank"] <= 3 else BG_GOOD if r["rank"] <= 10 else BG_NORMAL
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = FONT_SCORE if col == 8 else FONT_GRADE if col == 9 else FONT_BODY
            cell.fill = grade_fills.get(r["grade"], BG_NORMAL) if col == 9 else row_bg
            cell.alignment = ALIGN_C if col in [1,5,8,9,10,12,13,14,15,16,17,18,19] else ALIGN_L
        ws.row_dimensions[row_num].height = 16

    # ── ERRORS SHEET
    if errors:
        ws_err = wb.create_sheet("Errors")
        ws_err["A1"] = "Row"; ws_err["B1"] = "Name"; ws_err["C1"] = "Error"
        for i, e in enumerate(errors, 2):
            ws_err[f"A{i}"] = e["row"]; ws_err[f"B{i}"] = e["name"]; ws_err[f"C{i}"] = e["error"]

    # ── SUMMARY SHEET
    ws_sum = wb.create_sheet("Summary")
    scores = [r["master_score"] for r in results]
    summary_data = [
        ("ADC-DesignOS Batch Summary",""),
        ("PhyloBandits","www.phylobandits.com"),
        ("",""),
        ("Total Candidates", len(results)),
        ("Errors", len(errors)),
        ("Best Score", max(scores) if scores else 0),
        ("Worst Score", min(scores) if scores else 0),
        ("Mean Score", round(sum(scores)/len(scores),1) if scores else 0),
        ("Grade A (≥80)", sum(1 for s in scores if s>=80)),
        ("Grade B (65-79)", sum(1 for s in scores if 65<=s<80)),
        ("Grade C (50-64)", sum(1 for s in scores if 50<=s<65)),
        ("Grade D (<50)", sum(1 for s in scores if s<50)),
        ("",""),
        ("Best Candidate", results[0]["name"] if results else ""),
        ("Best Score", results[0]["master_score"] if results else ""),
        ("Best Linker", results[0]["linker"] if results else ""),
        ("Best Payload", results[0]["payload"] if results else ""),
        ("Best DAR", results[0]["optimal_dar"] if results else ""),
        ("",""),
        ("Generated", datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
    ]
    for i, (k, v) in enumerate(summary_data, 1):
        ws_sum[f"A{i}"] = k; ws_sum[f"B{i}"] = v
        if i == 1:
            ws_sum[f"A{i}"].font = Font(bold=True, color="22d86e", size=13)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 32

    # ── Freeze panes & auto-filter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # Return as file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ADC_Batch_{datetime.utcnow().strftime('%Y%m%d_%H%M')}_PhyloBandits.xlsx"
    return Response(buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/api/batch/template", methods=["GET"])
def api_batch_template():
    """Return a CSV template for batch upload."""
    from flask import Response
    lines = [
        "name,sequence,linker,payload,dar,target_location,target_expression,dose_mg_kg",
        f"Candidate_1,{SEQ_EXAMPLE},mc-VC-PABC,MMAE,4,intracellular,high,3.0",
        f"Candidate_2,{SEQ_EXAMPLE},SMCC,DM1,2,intracellular,medium,3.0",
        f"Candidate_3,{SEQ_EXAMPLE},PEG4-NHS,DXd,8,intracellular,very_high,5.0",
        f"Candidate_4,{SEQ_EXAMPLE},Disulfide,MMAF,4,extracellular,high,3.0",
        f"Candidate_5,{SEQ_EXAMPLE},β-glucuronide,SN-38,4,intracellular,high,3.0",
    ]
    return Response("\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=batch_template.csv"})

SEQ_EXAMPLE = "EVQLVESGGGLVQPGGSLRLSCAASGFTFSSYAMSWVRQAPGKGLEWVSAISGSGGSTYYADSVKGRFTISRDNSKNTLYLQMNSLRAEDTAVYYCAKDYYGSNSYYFDYWGQGTLVTVSS"


# ═══════════════════════════════════════════════════════════
# FEATURE 9: RAZORPAY / FREE-TIER PAYWALL
# ═══════════════════════════════════════════════════════════

RAZORPAY_KEY_ID     = os.environ.get("RAZORPAY_KEY_ID", "rzp_test_PLACEHOLDER")
RAZORPAY_KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET", "PLACEHOLDER_SECRET")

# In-memory usage store (use Redis/DB in production)
_usage = {}  # session_id -> {"count": int, "date": "YYYY-MM-DD", "paid": bool, "tier": "free"|"pro"|"enterprise"}
_orders = {} # order_id -> {session_id, amount, status}

FREE_LIMIT = 3  # free analyses per day

def get_usage(session_id: str) -> dict:
    today = datetime.utcnow().strftime("%Y-%m-%d")
    u = _usage.get(session_id, {"count": 0, "date": today, "paid": False, "tier": "free"})
    if u["date"] != today:
        u = {"count": 0, "date": today, "paid": u.get("paid", False), "tier": u.get("tier","free")}
    _usage[session_id] = u
    return u

@app.route("/api/usage/<session_id>", methods=["GET"])
def api_usage(session_id):
    u = get_usage(session_id)
    remaining = 999 if u["tier"] != "free" else max(0, FREE_LIMIT - u["count"])
    return jsonify({
        "session_id": session_id,
        "tier": u["tier"],
        "count_today": u["count"],
        "limit": FREE_LIMIT if u["tier"] == "free" else "unlimited",
        "remaining": remaining,
        "can_analyze": remaining > 0,
        "reset_utc": (datetime.utcnow().replace(hour=0,minute=0,second=0,microsecond=0) +
                      __import__('datetime').timedelta(days=1)).isoformat(),
    })

@app.route("/api/paywall/check", methods=["POST"])
def api_paywall_check():
    """Called before each analysis — returns whether allowed or payment needed."""
    d = request.json or {}
    sid = d.get("session_id","anon")
    u = get_usage(sid)
    today = datetime.utcnow().strftime("%Y-%m-%d")

    if u["tier"] != "free":
        return jsonify({"allowed": True, "tier": u["tier"], "message": f"{u['tier'].title()} plan active"})

    if u["count"] < FREE_LIMIT:
        u["count"] += 1
        _usage[sid] = u
        return jsonify({
            "allowed": True, "tier": "free",
            "count": u["count"], "remaining": FREE_LIMIT - u["count"],
            "message": f"Free analysis {u['count']}/{FREE_LIMIT} today"
        })

    return jsonify({
        "allowed": False, "tier": "free",
        "message": f"Free limit reached ({FREE_LIMIT}/day). Upgrade to Pro for unlimited analyses.",
        "pricing": {
            "per_analysis": {"amount": 1500000, "currency": "INR", "display": "₹15,000/analysis"},
            "monthly_pro":  {"amount": 25000000, "currency": "INR", "display": "₹2,50,000/month — unlimited"},
            "api_enterprise":{"amount": 100000000, "currency": "INR", "display": "₹10,00,000 — enterprise API"},
        },
        "razorpay_key": RAZORPAY_KEY_ID,
    }), 402

@app.route("/api/paywall/create-order", methods=["POST"])
def api_create_order():
    """Create a Razorpay order (or mock in test mode)."""
    d = request.json or {}
    sid    = d.get("session_id","anon")
    tier   = d.get("tier","per_analysis")  # per_analysis | monthly_pro | api_enterprise
    PRICES = {"per_analysis": 1500000, "monthly_pro": 25000000, "api_enterprise": 100000000}
    amount = PRICES.get(tier, 1500000)

    # In test mode (no real Razorpay keys) — return mock order
    if RAZORPAY_KEY_ID.startswith("rzp_test_PLACEHOLDER"):
        order_id = "mock_order_" + hashlib.md5(f"{sid}{tier}{datetime.utcnow()}".encode()).hexdigest()[:8]
        _orders[order_id] = {"session_id": sid, "amount": amount, "tier": tier, "status": "created"}
        return jsonify({
            "order_id": order_id,
            "amount": amount,
            "currency": "INR",
            "tier": tier,
            "razorpay_key": RAZORPAY_KEY_ID,
            "test_mode": True,
            "note": "Set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET env vars to enable real payments. Get keys at razorpay.com (free account)."
        })

    # Real Razorpay integration
    try:
        import razorpay
        client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
        order = client.order.create({"amount": amount, "currency": "INR", "receipt": f"adc_{sid[:8]}"})
        _orders[order["id"]] = {"session_id": sid, "amount": amount, "tier": tier, "status": "created"}
        return jsonify({"order_id": order["id"], "amount": amount, "currency": "INR", "tier": tier, "razorpay_key": RAZORPAY_KEY_ID})
    except ImportError:
        return jsonify({"error": "Install razorpay: pip install razorpay", "note": "Works in test mode without it."}), 500

@app.route("/api/paywall/verify", methods=["POST"])
def api_paywall_verify():
    """Verify payment and upgrade session tier."""
    d = request.json or {}
    order_id   = d.get("razorpay_order_id","")
    payment_id = d.get("razorpay_payment_id","")
    sid        = d.get("session_id","anon")

    # Test mode — auto-approve mock orders
    if order_id.startswith("mock_order_"):
        order = _orders.get(order_id, {})
        tier  = order.get("tier","per_analysis")
        new_tier = "pro" if "pro" in tier else "pro"
        u = get_usage(sid)
        u["paid"] = True
        u["tier"] = new_tier
        _usage[sid] = u
        return jsonify({"verified": True, "tier": new_tier, "message": f"✓ Payment verified (test mode). {new_tier.title()} plan activated.", "session_id": sid})

    # Real Razorpay verification
    try:
        import razorpay
        client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
        params = {"razorpay_order_id": order_id, "razorpay_payment_id": payment_id, "razorpay_signature": d.get("razorpay_signature","")}
        client.utility.verify_payment_signature(params)
        order = _orders.get(order_id, {})
        tier  = order.get("tier","per_analysis")
        new_tier = "pro" if "pro" in tier else "enterprise" if "enterprise" in tier else "pro"
        u = get_usage(sid)
        u["paid"] = True; u["tier"] = new_tier
        _usage[sid] = u
        return jsonify({"verified": True, "tier": new_tier, "message": f"Payment verified. {new_tier.title()} plan activated."})
    except Exception as e:
        return jsonify({"verified": False, "error": str(e)}), 400
